#!/usr/bin/env python3
"""
Xiaohongshu metric guard.

Purpose:
1) Distinguish account-level and content-level click definitions.
2) Warn when key fields are missing before importing into markdown tables.
3) Optionally validate the target markdown contains required label names.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


REQUIRED_MD_LABELS = [
    "平台封面点击率（账号口径）",
    "观看/曝光比（自算口径）",
    "内容点击率（单条口径）",
]


@dataclass
class LocatedMetric:
    value: float
    sheet: str
    cell: str
    raw: str


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().replace(" ", "").replace("\n", "")


def parse_numeric(value: object) -> float | None:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip().replace(",", "").replace("，", "")
    if not s or s in {"-", "--", "—"}:
        return None

    multiplier = 1.0
    if "万" in s:
        multiplier = 10000.0
    elif "亿" in s:
        multiplier = 100000000.0

    percent = "%" in s
    match = re.search(r"-?\d+(?:\.\d+)?", s)
    if not match:
        return None
    number = float(match.group(0)) * multiplier

    if percent:
        return number / 100.0
    return number


def parse_rate(value: object) -> float | None:
    number = parse_numeric(value)
    if number is None:
        return None
    if number > 1.0:
        return number / 100.0
    return number


def find_metric(
    workbook,
    keywords: Iterable[str],
    parser,
    preferred_sheet_keywords: Iterable[str] | None = None,
    max_rows: int = 160,
    max_cols: int = 26,
) -> LocatedMetric | None:
    offsets = [(0, 1), (0, 2), (0, 3), (1, 0), (2, 0), (1, 1)]
    preferred_sheet_keywords = list(preferred_sheet_keywords or [])

    def iter_sheets():
        if not preferred_sheet_keywords:
            yield from workbook.worksheets
            return
        preferred = []
        others = []
        for sh in workbook.worksheets:
            if any(k in sh.title for k in preferred_sheet_keywords):
                preferred.append(sh)
            else:
                others.append(sh)
        yield from preferred
        yield from others

    for sheet in iter_sheets():
        limit_row = min(sheet.max_row, max_rows)
        limit_col = min(sheet.max_column, max_cols)
        for r in range(1, limit_row + 1):
            for c in range(1, limit_col + 1):
                cell = sheet.cell(row=r, column=c)
                label = normalize_text(cell.value)
                if not label:
                    continue
                if not any(k in label for k in keywords):
                    continue
                for dr, dc in offsets:
                    rr = r + dr
                    cc = c + dc
                    if rr < 1 or cc < 1 or rr > sheet.max_row or cc > sheet.max_column:
                        continue
                    candidate = sheet.cell(row=rr, column=cc)
                    parsed = parser(candidate.value)
                    if parsed is None:
                        continue
                    return LocatedMetric(
                        value=parsed,
                        sheet=sheet.title,
                        cell=candidate.coordinate,
                        raw=str(candidate.value),
                    )
    return None


def find_content_sheet(workbook):
    preferred = ["内容数据", "笔记数据", "内容明细", "作品数据", "内容"]
    for name in preferred:
        if name in workbook.sheetnames:
            return workbook[name]
    # fallback: choose sheet with most rows
    return max(workbook.worksheets, key=lambda s: s.max_row)


def locate_header_row(sheet, max_scan_rows: int = 15) -> tuple[int, dict[str, int]] | None:
    def has_any(text: str, keys: Iterable[str]) -> bool:
        return any(k in text for k in keys)

    for r in range(1, min(sheet.max_row, max_scan_rows) + 1):
        values = [normalize_text(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1)]
        mapping: dict[str, int] = {}
        for idx, text in enumerate(values, start=1):
            if not text:
                continue
            if "title" not in mapping and has_any(text, ["标题", "笔记标题", "内容标题"]):
                mapping["title"] = idx
            if "exposure" not in mapping and has_any(text, ["曝光", "曝光量", "曝光次数"]):
                mapping["exposure"] = idx
            if "watch" not in mapping and has_any(text, ["观看", "观看量", "播放量", "播放"]):
                mapping["watch"] = idx
            if "click" not in mapping and has_any(text, ["点击率", "封面点击率"]):
                mapping["click"] = idx
        if {"exposure", "watch"} <= set(mapping):
            return r, mapping
    return None


def analyze_content_rows(sheet):
    head = locate_header_row(sheet)
    if not head:
        return {
            "rows": 0,
            "missing_click": [],
            "extreme_delta": [],
            "header_found": False,
            "has_click_col": False,
        }

    header_row, mapping = head
    title_col = mapping.get("title")
    exposure_col = mapping["exposure"]
    watch_col = mapping["watch"]
    click_col = mapping.get("click")

    rows = 0
    missing_click: list[str] = []
    extreme_delta: list[str] = []
    empty_streak = 0

    for r in range(header_row + 1, sheet.max_row + 1):
        exposure = parse_numeric(sheet.cell(r, exposure_col).value)
        watch = parse_numeric(sheet.cell(r, watch_col).value)
        click = parse_rate(sheet.cell(r, click_col).value) if click_col else None
        title = normalize_text(sheet.cell(r, title_col).value) if title_col else ""
        display_title = title or f"第{r}行"

        if exposure is None and watch is None and click is None and not title:
            empty_streak += 1
            if empty_streak >= 10:
                break
            continue
        empty_streak = 0

        if exposure is None and watch is None:
            continue

        rows += 1
        if click_col and click is None:
            missing_click.append(display_title)

        if exposure and exposure > 0 and click is not None:
            ratio = watch / exposure if watch is not None else None
            if ratio is not None and abs(click - ratio) >= 0.2:
                extreme_delta.append(
                    f"{display_title}（内容点击率{click:.1%} vs 观看/曝光{ratio:.1%}）"
                )

    return {
        "rows": rows,
        "missing_click": missing_click,
        "extreme_delta": extreme_delta,
        "header_found": True,
        "has_click_col": click_col is not None,
    }


def check_markdown_labels(md_path: Path) -> list[str]:
    if not md_path.exists():
        return [f"目标文件不存在：{md_path}"]
    text = md_path.read_text(encoding="utf-8")
    missing = [label for label in REQUIRED_MD_LABELS if label not in text]
    return missing


def percent(value: float | None) -> str:
    if value is None:
        return "N/A"
    return f"{value:.1%}"


def build_report(
    source_file: Path,
    account_ctr: LocatedMetric | None,
    exposure: LocatedMetric | None,
    watch: LocatedMetric | None,
    content_stats: dict,
    md_path: Path | None,
    md_missing: list[str] | None,
) -> tuple[str, bool]:
    lines: list[str] = []
    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    lines.append(f"# 小红书导入口径检查报告")
    lines.append(f"- 检查时间：{now}")
    lines.append(f"- 数据文件：{source_file}")
    lines.append("")

    lines.append("## 1) 账号口径 vs 自算口径")
    lines.append(
        f"- 平台封面点击率（账号口径）：{percent(account_ctr.value) if account_ctr else '未识别'}"
    )
    if account_ctr:
        lines.append(f"  - 来源：{account_ctr.sheet}!{account_ctr.cell}（原值：{account_ctr.raw}）")

    ratio = None
    if exposure and watch and exposure.value > 0:
        ratio = watch.value / exposure.value
    lines.append(f"- 观看/曝光比（自算口径）：{percent(ratio)}")
    if exposure and watch:
        lines.append(
            f"  - 计算：观看 {watch.value:,.0f} / 曝光 {exposure.value:,.0f}"
        )
    lines.append("- 说明：这两个指标定义不同，禁止互相替代。")

    needs_attention = False
    if account_ctr is None:
        lines.append("- 风险：未识别到账号层封面点击率，导入时容易误用。")
        needs_attention = True
    if ratio is None:
        lines.append("- 风险：未识别到账号层曝光/观看，无法计算自算口径。")
        needs_attention = True
    if account_ctr and ratio is not None:
        delta = abs(account_ctr.value - ratio)
        lines.append(f"- 口径差值：{delta:.1%}")
        if delta >= 0.2:
            lines.append("- 提示：差值较大是常见情况，通常代表账号口径与内容口径并存。")

    lines.append("")
    lines.append("## 2) 单条内容口径检查")
    if not content_stats["header_found"]:
        lines.append("- 风险：未识别到内容表头（曝光/观看），请手动核对 sheet 结构。")
        needs_attention = True
    else:
        lines.append(f"- 识别到内容行数：{content_stats['rows']}")
        lines.append(
            f"- 单条点击率字段：{'已识别' if content_stats['has_click_col'] else '未识别'}"
        )
        if not content_stats["has_click_col"]:
            lines.append("- 风险：内容表没有点击率字段，导入到明细表时请勿写成账号口径。")
            needs_attention = True
        if content_stats["missing_click"]:
            lines.append(f"- 缺失单条点击率：{len(content_stats['missing_click'])} 条")
            for name in content_stats["missing_click"][:10]:
                lines.append(f"  - {name}")
            needs_attention = True
        if content_stats["extreme_delta"]:
            lines.append(
                f"- 单条口径差异较大（内容点击率 vs 观看/曝光）：{len(content_stats['extreme_delta'])} 条"
            )
            for name in content_stats["extreme_delta"][:8]:
                lines.append(f"  - {name}")

    lines.append("")
    lines.append("## 3) 目标 Markdown 标签检查")
    if md_path is None:
        lines.append("- 跳过（未提供目标 markdown 文件）")
    else:
        lines.append(f"- 目标文件：{md_path}")
        if md_missing:
            lines.append("- 风险：缺少以下标签，未来导入容易混口径：")
            for item in md_missing:
                lines.append(f"  - {item}")
            needs_attention = True
        else:
            lines.append("- 通过：关键标签齐全。")

    lines.append("")
    lines.append("## 导入前结论")
    lines.append(
        "- " + ("有风险项，建议先修正再导入。" if needs_attention else "检查通过，可继续导入。")
    )
    return "\n".join(lines) + "\n", needs_attention


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="检查小红书数据导入时的指标口径混用风险")
    parser.add_argument("xlsx", type=Path, help="小红书导出的 xlsx 文件路径")
    parser.add_argument(
        "--md",
        type=Path,
        help="要检查标签是否齐全的 markdown 文件（如 内容数据表.md）",
    )
    parser.add_argument("--out", type=Path, help="将检查报告写入指定文件（.md）")
    parser.add_argument(
        "--strict",
        action="store_true",
        help="发现风险项时返回非零退出码，便于流程卡口",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.xlsx.exists():
        print(f"文件不存在：{args.xlsx}", file=sys.stderr)
        return 2

    wb = load_workbook(args.xlsx, data_only=True)
    account_sheet_hints = ["账号", "概览", "总体", "总览"]
    account_ctr = find_metric(
        wb, ["封面点击率"], parse_rate, preferred_sheet_keywords=account_sheet_hints
    )
    exposure = find_metric(
        wb, ["曝光"], parse_numeric, preferred_sheet_keywords=account_sheet_hints
    )
    watch = find_metric(
        wb, ["观看"], parse_numeric, preferred_sheet_keywords=account_sheet_hints
    )
    content_sheet = find_content_sheet(wb)
    content_stats = analyze_content_rows(content_sheet)
    md_missing = check_markdown_labels(args.md) if args.md else None

    report, has_risk = build_report(
        source_file=args.xlsx,
        account_ctr=account_ctr,
        exposure=exposure,
        watch=watch,
        content_stats=content_stats,
        md_path=args.md,
        md_missing=md_missing,
    )

    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(report, encoding="utf-8")
        print(f"报告已写入：{args.out}")
    else:
        print(report)

    if args.strict and has_risk:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
